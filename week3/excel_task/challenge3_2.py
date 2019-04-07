# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
    '''
    open workbook xlsx
    select sheet range -student
    active worksheet -student
    loop row and column
    read value from cell
    fill value to student-list for create,course,amount
    
    active worksheet -time
    loop row and column
    read value from cell
    fill value to time-list for create,course,study

    create list combine create,course,amount,study
    fill value based on course name

    create new sheet - combine
    loop row and column
    write list value to cell

    '''

    wb=load_workbook(filename='courses.xlsx')
    ws1=wb['students']
    ws2=wb['time']
    list_student=[]
    for x in range(2,486):
        dict_student={}
        dict_student['create']=(ws1.cell(row=x,column=1)).value
        dict_student['course']=(ws1.cell(row=x,column=2)).value
        dict_student['amount']=(ws1.cell(row=x,column=3)).value
        list_student.append(dict_student)

    list_time=[]
    for x in range(2,486):
        dict_time={}
        dict_time['create']=(ws2.cell(row=x,column=1)).value
        dict_time['course']=(ws2.cell(row=x,column=2)).value
        dict_time['study']=(ws2.cell(row=x,column=3)).value
        list_time.append(dict_time)

    list_combine=[]
    for x in range(0,484):
        study=0
        dict_combine={}
        dict_student=list_student[x]
        dict_combine['create']=dict_student['create']
        dict_combine['course']=dict_student['course']
        dict_combine['amount']=dict_student['amount']
        for item in list_time:
            if item['course']==dict_student['course']:
                dict_combine['study']=item['study']
                study=item['study']
                break
        dict_combine['study']=study
        list_combine.append(dict_combine)

    list_combine_time=sorted(list_combine,key=lambda k:k['create'])
    #for index,item in enumerate(list_combine_time):
        #print(index,item)

    ws3=wb.create_sheet(title="combine")
    ws3.cell(row=1,column=1).value="created time"    
    ws3.cell(row=1,column=2).value="course name"    
    ws3.cell(row=1,column=3).value="amount student"    
    ws3.cell(row=1,column=4).value="study time"    
    for x in range(2,486):
        dict_combine=list_combine_time[x-2]
        ws3.cell(row=x,column=1).value=dict_combine['create']    
        ws3.cell(row=x,column=2).value=dict_combine['course'] 
        ws3.cell(row=x,column=3).value=dict_combine['amount'] 
        ws3.cell(row=x,column=4).value=dict_combine['study'] 

    wb.save(filename='combine.xlsx')
    return list_combine_time
        
def split(list_combine_time):
    
    wb_2013=Workbook()
    ws_2013=wb_2013.active

    wb_2014=Workbook()
    ws_2014=wb_2014.active

    wb_2015=Workbook()
    ws_2015=wb_2015.active

    wb_2016=Workbook()
    ws_2016=wb_2016.active
    
    year=2013
    for worksheet in [ws_2013,ws_2014,ws_2015,ws_2016]:
        worksheet.title=str(year)
        worksheet.cell(row=1,column=1).value="created time"    
        worksheet.cell(row=1,column=2).value="course name"    
        worksheet.cell(row=1,column=3).value="amount student"    
        worksheet.cell(row=1,column=4).value="study time"    
        year+=1
        index_2013,index_2014,index_2015,index_2016=2,2,2,2
    for x in range(2,486):
        dict_combine=list_combine_time[x-2]
        if dict_combine['create'].year==2013:
            ws_2013.cell(row=index_2013,column=1).value=dict_combine['create']    
            ws_2013.cell(row=index_2013,column=2).value=dict_combine['course'] 
            ws_2013.cell(row=index_2013,column=3).value=dict_combine['amount'] 
            ws_2013.cell(row=index_2013,column=4).value=dict_combine['study'] 
            index_2013+=1
        elif dict_combine['create'].year==2014:
            ws_2014.cell(row=index_2014,column=1).value=dict_combine['create']    
            ws_2014.cell(row=index_2014,column=2).value=dict_combine['course'] 
            ws_2014.cell(row=index_2014,column=3).value=dict_combine['amount'] 
            ws_2014.cell(row=index_2014,column=4).value=dict_combine['study'] 
            index_2014+=1
        elif dict_combine['create'].year==2015:
            ws_2015.cell(row=index_2015,column=1).value=dict_combine['create']    
            ws_2015.cell(row=index_2015,column=2).value=dict_combine['course'] 
            ws_2015.cell(row=index_2015,column=3).value=dict_combine['amount'] 
            ws_2015.cell(row=index_2015,column=4).value=dict_combine['study'] 
            index_2015+=1
        elif dict_combine['create'].year==2016:
            ws_2016.cell(row=index_2016,column=1).value=dict_combine['create']    
            ws_2016.cell(row=index_2016,column=2).value=dict_combine['course'] 
            ws_2016.cell(row=index_2016,column=3).value=dict_combine['amount'] 
            ws_2016.cell(row=index_2016,column=4).value=dict_combine['study'] 
            index_2016+=1
        else:
            print("wow , why here.year should be between 2013 to 2016")
    wb_2013.save(filename='2013.xlsx')
    wb_2014.save(filename='2014.xlsx')
    wb_2015.save(filename='2015.xlsx')
    wb_2016.save(filename='2016.xlsx')

if __name__ == '__main__':
    list_split=combine()
    split(list_split)
