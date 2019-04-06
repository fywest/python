from openpyxl import load_workbook
from openpyxl import workbook
import datetime
wb=load_workbook(filename='courses.xlsx')
ws=wb['students']
list_student=[]
for x in range(1,11):
    dict_student={}
    dict_student['create']=(ws.cell(row=x,column=1)).value
    dict_student['course']=(ws.cell(row=x,column=2)).value
    dict_student['study']=(ws.cell(row=x,column=3)).value
    print(dict_student)   
    list_student.append(dict_student)
for item in list_student:
    print(item)
    
    
        
