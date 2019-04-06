# coding: utf-8
get_ipython().run_line_magic('load', 'excel_1.py')
# %load excel_1.py
from openpyxl import load_workbook
from openpyxl import workbook
import datetime
wb=load_workbook(filename='courses.xlsx')
sheet_ranges=wb['students']
print(sheet_ranges)
ws=wb.get_active_sheet('students')
ws=wb.activate()
ws=wb.active
ws.title
print(wb.sheetnames)

for x in range(1,486):
    str_text=' '
    for y in range(1,2):
        text=ws.cell(row=x,column=y)
        print(text.value)
        #str_text+=str(text.value)+'\t '
    #print(str_text)
    
        
